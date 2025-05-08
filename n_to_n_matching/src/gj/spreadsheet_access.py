#!/usr/bin/env python

# Copyright 2024 Isaac Saito.
#
# Licensed under the Apache License, Version 2.0 (the "License");
# you may not use this file except in compliance with the License.
# You may obtain a copy of the License at
#
#     http://www.apache.org/licenses/LICENSE-2.0
#
# Unless required by applicable law or agreed to in writing, software
# distributed under the License is distributed on an "AS IS" BASIS,
# WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND, either express or implied.
# See the License for the specific language governing permissions and
# limitations under the License.

from abc import ABC, abstractmethod
import openpyxl as pyxl
from openpyxl.cell.cell import Cell as pyxl_Cell
from typing import Dict, List

from gj.grade_class import GjGrade, GradeUtil
from gj.responsibility import Responsibility, ResponsibilityLevel
from gj.role import Role, Roles_Definition
from gj.util import GjUtil
from n_to_n_matching.person_player import PersonBank, PersonPlayer
from n_to_n_matching.spreadsheet_access import SpreadsheetCell, SpreadsheetRow


class GjRowEntity:
    COLTITLE_ID_IN_SHEET = "id_in_sheet"
    COLTITLE_GRADE_CLASS = "grade_class"
    COLTITLE_STUDENT_NAME = "person_name"
    COLTITLE_GUARDIAN_NAME = "guardian_personname"
    COLTITLE_PHONE_EMERGENCY = "phone_emergency"
    COLTITLE_EMAIL_EMERGENCY = "email_emergency"
    COLTITLE_SIBLING_2_CLASS = "sibling_2_class"
    COLTITLE_SIBLING_2_PERSONNAME = "sibling_2_personname"
    COLTITLE_SIBLING_3_CLASS = "sibling_3_class"
    COLTITLE_SIBLING_3_PERSONNAME = "sibling_3_personname"
    COLTITLE_SIBLING_4_CLASS = "sibling_4_class"
    COLTITLE_DATE_ASSIGNED_TOSHO = "date_assigned_tosho"
    COLTITLE_DATE_ASSIGNED_HOKEN = "date_assigned_hoken"
    COLTITLE_DATE_ASSIGNED_PATROL = "date_assigned_patrol"
    COLTITLE_COMMENT = "comment"
    COLTITLE_TRANSFERED_DATE = "transfered_on"
    COLTITLE_TERMINATE_DATE = "terminate_on"
    COLTITLE_EXEMPTED_BY = "exempted_on"
    COLTITLE_SELECTED_AS = "selected_as"
    COLTITLE_PHONENUM_REGISTERED = "phone_registered"
    COLTITLE_EMAIL_REGISTERED = "email_registered"

    COL_TITLE_IDS_20250503 = {
        2: COLTITLE_ID_IN_SHEET,           # "No"
        3: COLTITLE_GRADE_CLASS,           # "学年組"
        4: COLTITLE_STUDENT_NAME,           # "氏名"
        5: COLTITLE_GUARDIAN_NAME,   # "保護者名"
        6: COLTITLE_PHONE_EMERGENCY,       # "当番用TEL"
        7: COLTITLE_EMAIL_EMERGENCY,       # "当番用メール"
        8: COLTITLE_SIBLING_2_CLASS,       # "兄姉２"
        9: COLTITLE_SIBLING_2_PERSONNAME,  # "兄姉氏名"
        10: COLTITLE_SIBLING_3_CLASS,       # "兄姉３"
        11: COLTITLE_SIBLING_3_PERSONNAME,  # "兄姉氏名"
        12: COLTITLE_SIBLING_4_CLASS,       # "兄姉４"
        14: COLTITLE_DATE_ASSIGNED_TOSHO,   # "図書" NOTE: From  15 (図書) to 20 (退学) aren't verified yet as of 20240829.
        15: COLTITLE_DATE_ASSIGNED_HOKEN,   # "保健"
        16: COLTITLE_DATE_ASSIGNED_PATROL,  # "パトロール"
        17: COLTITLE_COMMENT,               # "備考"
        18: COLTITLE_TRANSFERED_DATE,         # "編入"
        19: COLTITLE_TERMINATE_DATE,          # "退学"
        23: COLTITLE_EXEMPTED_BY,           # "免除対象"
        24: COLTITLE_SELECTED_AS,           # "選出". As of 20240825 no practical usage of this field is found.
        25: COLTITLE_PHONENUM_REGISTERED,      # "事務局登録TEL"
        26: COLTITLE_EMAIL_REGISTERED,      # "クラス登録メール"
    }

    COL_TITLE_IDS_202407 = {
        2: COLTITLE_ID_IN_SHEET,           # "No"
        3: COLTITLE_GRADE_CLASS,           # "学年組"
        5: COLTITLE_STUDENT_NAME,           # "氏名"
        6: COLTITLE_GUARDIAN_NAME,   # "保護者名"
        7: COLTITLE_PHONE_EMERGENCY,       # "当番用TEL"
        8: COLTITLE_EMAIL_EMERGENCY,       # "当番用メール"
        9: COLTITLE_SIBLING_2_CLASS,       # "兄姉２"
        10: COLTITLE_SIBLING_2_PERSONNAME,  # "兄姉氏名"
        11: COLTITLE_SIBLING_3_CLASS,       # "兄姉３"
        12: COLTITLE_SIBLING_3_PERSONNAME,  # "兄姉氏名"
        13: COLTITLE_SIBLING_4_CLASS,       # "兄姉４"
        15: COLTITLE_DATE_ASSIGNED_TOSHO,   # "図書" NOTE: From  15 (図書) to 20 (退学) aren't verified yet as of 20240829.
        16: COLTITLE_DATE_ASSIGNED_HOKEN,   # "保健"
        17: COLTITLE_DATE_ASSIGNED_PATROL,  # "パトロール"
        18: COLTITLE_COMMENT,               # "備考"
        19: COLTITLE_TRANSFERED_DATE,         # "編入"
        20: COLTITLE_TERMINATE_DATE,          # "退学"
        24: COLTITLE_EXEMPTED_BY,           # "免除対象"
        25: COLTITLE_SELECTED_AS,           # "選出". As of 20240825 no practical usage of this field is found.
        26: COLTITLE_PHONENUM_REGISTERED,      # "事務局登録TEL"
        27: COLTITLE_EMAIL_REGISTERED,      # "クラス登録メール"
    }

    def __init__(self, row: SpreadsheetRow, row_spec=COL_TITLE_IDS_20250503, logger_obj=None):
        if type(row) != SpreadsheetRow:
            raise ValueError(f"'row' object must be the type of SpreadsheetRow. Instead, {type(row)} was passed.")
        self._logger = GjUtil.get_logger(__name__, logger_obj)
        if not row:
            self._logger.warning(f"'row' object is empty. Finishing creating an object without filling with the values.")
            pass
        else:
            self._raw_row = row

        self._is_title_row = True if row.is_title_row else False
        self._row_spec = row_spec
        self._gj_row = self.parse_xls_row(row, self._row_spec)
    
    def parse_xls_row(self, row: SpreadsheetRow, col_title_definition=COL_TITLE_IDS_20250503) -> List[SpreadsheetCell]:
        """
        @description: Evaluate each cell from the row object, find a cell that matches the column ID, then sets the cell's value.
        @type row: SpreadsheetRow
        @type col_title_definition: dict
        @return: row object filled with the values in the corresponding column index.
        """
        ss_row = []
        for cell in row.cells:
            #if (cell.column == 3) and (not ss_cell.value):
                # TODO Specifying columnd=3 by hardcoding is too adhoc.
                #self._logger.debug(f"At {cell.row=}, Grade-Class is empty. Skipping this row.")
                #return None

            if (cell.row % 7 == 0) and (cell.row < 255) and (cell.value):  # This number is very adhoc
                self._logger.info(f"{cell.row=}-{cell.column=}, {cell.value=}")
            cell_title = ""
            try:
                cell_title = col_title_definition[cell.column]
            except KeyError as e:
                self._logger.debug(f"Skipping this column. {cell.column=} is either not used in the input file, or its definition not present in title definition `col_title_definition`.")                
            # TODO Roles need special treatment; In the spreadsheet, roles are distributed in multiple columns and value of the cells in each col
            #   are meant to be the date the person is assigned to each role. While this design makes sense from spreadsheet user's usecase,
            #   it doesn't provide convenience for this tool's usecases.
            #   To mitigate, add a special cell to the row object that conveys the type of role.
            ss_cell = SpreadsheetCell(cell_obj=cell, title_val=cell_title)
            # TODO Looks like the usage of this ss_row list assumes the order of elements are fixed, however,
            # the current implementation doesn't guarantee the order. Needs a better impl.

            #try:
                #self._logger.debug(f"{ss_cell=}")  ## 20250503 This logger call fails
            #except AttributeError as e:
            #    self._logger.error(f"Error happened. Continuing though.\n\tError content: {str(e)}")

            ss_row.append(ss_cell)
        return ss_row

    def _get_key_byval(self, dct: Dict[int, str], value: str) -> List[int]:
        """
        @summary Return the corresponding key(s) from the dict that has the given `value`.
        @see https://stackoverflow.com/a/49353279/577001
        """
        return [key for key in dct if (dct[key] == value)]

    def _get_value_from_gj_row(self, col_id_str: str) -> str:
        """
        @param col_id_str: The title of the column in the given spreadsheet.
          Internally look up the row ID database that must already be passed upon initializing the class.
        """
        col_ids = self._get_key_byval(self._row_spec, col_id_str)
        if 1 < len(col_ids):
            raise ValueError(f"In the usecase here, '{col_id_str=}' is supposed to be 1 or 0.")
        col_id = col_ids[0]
        for cell in self._raw_row.cells:  # type(cell) == openpyxl.cell.cell.Cell
            self._logger.debug(f"{cell.column=}, {col_id=}")
            if cell.column == col_id:
                self._logger.debug(f"{cell.column=} == {col_id=}. {cell.value=}")
                return cell.value

    @property
    def row_id(self):
        return self._raw_row.row_id

    @property
    def exempted_on(self):
        return self._get_value_from_gj_row(self.COLTITLE_EXEMPTED_BY)

    @property
    def grade_class(self):
        return self._get_value_from_gj_row(self.COLTITLE_GRADE_CLASS)

    @property
    def id_in_sheet(self):
        return self._get_value_from_gj_row(self.COLTITLE_ID_IN_SHEET)

    @property
    def person_name(self):
        name = self._get_value_from_gj_row(self.COLTITLE_STUDENT_NAME)
        if not name:
            raise ValueError("person_name is empty.")
        return name

    @property
    def phone_emergency(self):
        return self._get_value_from_gj_row(self.COLTITLE_PHONE_EMERGENCY)

    @property
    def email_emergency(self):
        return self._get_value_from_gj_row(self.COLTITLE_EMAIL_EMERGENCY)


class GjToubanAccess:
    """
    @todo: Spreadsheet format is tied to .xls as of 2024/08 but no guarantee to stick with it in the future (so better keep that in mind when making design decisions).
    """
    SUFFIX_MASTERSHEET = "マスター"
    _MASTERSHEET_2024 = "2024当番マスター"
    # ID of the column in a spreadsheet that shows the responsibilitys that the exemption rule is to eb applied for certain assignment.
    COL_ROW_EXEMPT = "X"
    NAME_TOSHOIIN = "図書委員"
    LIST_AVAILABLE_TARGET = [NAME_TOSHOIIN]

    def __init__(self, logger_obj=None):
        self._logger = GjUtil.get_logger(__name__, logger_obj)
        self._touban_master_sheet = None

    @staticmethod
    def get_a_sheet_by_name(workbook, sheet_name):
        """
        @type workbook: openpyxl.workbook.workbook.Workbook
        @param sheet_name: The full name of the sheet to obtain.
        @rtype: Worksheet
        """
        for sheet in workbook:
            if sheet.title == sheet_name:
                return sheet
        raise LookupError(f"Requested sheet '{sheet_name}' not found in the given workbook obj.")

    def get_a_spread_sheet(self, path_xls, sheet_name=_MASTERSHEET_2024):
        # 'data_only=True' is needed in order to read a value from each cell, not the macro formula.
        # See https://stackoverflow.com/a/35624928/577001
        wb = pyxl.load_workbook(path_xls, data_only=True)
        sheet = self.get_a_sheet_by_name(wb, sheet_name)
        if not sheet:
            raise ValueError("In the workbook '{}', no sheet found that has the name '{}'".format(
                wb.title, sheet_name))
        return sheet

    @staticmethod
    def get_a_sheet(workbook, suffix_master_file):
        """
        @deprecated: Picking up a sheet just by a suffix of the sheet name may not be robust.
        @type workbook: openpyxl.workbook.workbook.Workbook
        @rtype: Worksheet
        """
        for sheet in workbook:
            if sheet.title.endswith(suffix_master_file):
                return sheet

    @staticmethod
    def first_empty_row(sheet: pyxl.worksheet.worksheet) -> int:
        """
        @description: Find the first empty row in the given sheet.
        """
        # Iterate through each row in the sheet and check if all cells are None    
        for row_num, row in enumerate(sheet.iter_rows(min_row=1, max_row=sheet.max_row), start=1):
            if all(cell.value is None for cell in row):
                return row_num
        return sheet.max_row + 1 # If no empty row is found, return the next row number

    def get_touban_master_sheet(self, path_xls, sheet_name=_MASTERSHEET_2024):
        if not path_xls:
            raise ValueError("Either/Both args 'path_xls' and/or 'sheet_struct' is empty.")
        sheet = self.get_a_spread_sheet(path_xls, sheet_name)
        return sheet

    def get_candidates(self, sheet, key_target):
        """
        @param key_target: E.g. NAME_TOSHOIIN
        @rtype: 1) [[cell]], 2) [int]
        """
        if key_target not in self.LIST_AVAILABLE_TARGET:
            raise ValueError("The passed key_target='{}' is not present in the available targets '{}'.".format(
                key_target, self.LIST_AVAILABLE_TARGET))
        
        rows_matched = []
        # ID of the rows that meet the search criteria
        row_ids = []
        for cell in sheet[self.COL_ROW_EXEMPT]:
            if cell.value == key_target:
                row_ids.append(cell.row)  # 'row' is not intuitive way to get row number but it works.
        self._logger.info(f"row_ids: {row_ids}")    
        # Haven't found openpyxl API way of knowing the row number,
        # so maually figuring here...
        for row_id in row_ids:
            row = sheet[row_id]
            # Verify the row ID matches with what we want.
            if row[0].row == row_id:  # Again, 'row' returning row ID is unintuitive, but this is the way..
                rows_matched.append(sheet[row_id])
        self._logger.debug(f"Rows matched: {rows_matched}")
        return rows_matched, row_ids

    def gj_xls_to_personobj(self, path_to_xls: str, sheet_name: str, title_row=3, row_spec=GjRowEntity.COL_TITLE_IDS_20250503) -> PersonBank:
        """
        @description Convert GJLS' .xls specific format to the format this package can handle.

            Assumption for the spreadsheet format:
            - Titles of cells are defined in a single row.
            - Rows before the title row does not contain any info that needs to be taken into consideration for creating person info.

        @param title_row: The row where the values represent the type of the info that the cells under each column carries.
          As of 20240827, this title row must be a single row (i.e. Cases where titles are written in multiple rows are not yet supported).
        """
        persons = []
        # Read .xls file into a Python objects
        rows_xls_obj = self.get_touban_master_sheet(path_to_xls, sheet_name=sheet_name)
        # Each row should obtain the ID number from a cell in each row in the spreadsheet,
        # but how reliably maintained the ID in the spreadsheet is unknown. So here
        # maintaining ID as well. This is just a backup.
        _row_count = 0
        # Parse each row object, create 'PersonPlayer' object per each person.        

        _max_row_id = self.first_empty_row(rows_xls_obj)
        self._logger.info(f"{rows_xls_obj.max_row=}, {_max_row_id}")
        for row_pyxl in rows_xls_obj:
            if _max_row_id < _row_count:
                break
            # If the row is title, set title flag.
            is_title_row = True if row_pyxl[0].row == title_row else False

            # If the row is earlier thatn title row, do not look for person info.
            if row_pyxl[0].row <= title_row:
                continue

            row = GjRowEntity(SpreadsheetRow(row_pyxl, is_title_row), row_spec, self._logger)
            _row_count += 1
            # Identify GJ role(s), and deduce the responsibility from the role(s).
            a_role = Role(row.exempted_on)
            _responsibility = None
            try:
                # TODO Assign role in addition to responsibility, for Tosho, Patrol.
                #_responsibility_id = self.match_responsibility(a_role)
                _responsibility = GjUtil.corresponding_responsibility(a_role)
            except ValueError as e:
                self._logger.error(f"Column #{row.row_id}. Skipping as an unknown error occurred. {str(e)}")
                continue

            if row.id_in_sheet:
               _family_id_in_sheet = int(row.id_in_sheet)
            else:
                _family_id_in_sheet = _row_count
            self._logger.debug(f"{_family_id_in_sheet=}")

            _student_fullname = ""
            try:
                _student_fullname = row.person_name
            except ValueError as e:
                self._logger.warning(f"{_student_fullname=} empty at {_row_count=}. Likely empty row. Skipping.")
                continue
  
            # TODO Not fully sure if 'exempted_on' is the correct selection.
            #responsibility = GjUtil.gen_responsibility(row.exempted_on)
            if not row.grade_class:
                self._logger.warning(f"Grade/Class is empty at {_row_count=}.")
            self._logger.debug(f"Grade: {row.grade_class}")
            person = PersonPlayer(
                id =_family_id_in_sheet,
                name =_student_fullname,
                email_addr = row.email_emergency,
                phone_num = row.phone_emergency,
                grade_class = GradeUtil.find_grade(row.grade_class),  # For Grade/Class there's a designated Python class so match the input to one.
                roles=[a_role],
                # 2024/08 'children_ids' attribute was originally created without the knowledge of how students/guardians are 
                # grouped into a family info. Now that it's more known, 'children_ids' doesn't seem to be needed, hence
                # setting 'None' here
                children_ids = None,
                responsibilities=[_responsibility],
            )
            self._logger.debug(f"person ID: {person.id}, name: {person.name}")
            persons.append(person)
        self._logger.debug(f"Persons: {persons}, size of persons: {len(persons)}")
        if 0 == len(persons):
            raise RuntimeError(f"No person found, or at least not detected, from the gievn spreadsheet ({path_to_xls=}).")
        return PersonBank(persons)

    @abstractmethod
    def match_responsibility(self, a_role_id: Roles_Definition=""):
        """
        @deprecated: Potentially this method may no longer needed, replaced by GjUtil.corresponding_responsibility().
        """
        raise NotImplementedError()


class GjToubanAccess2024(GjToubanAccess):
    def match_responsibility(self, a_role_id: Roles_Definition="") -> Responsibility:
        """
        @description: (Overriding abstract method)
        @raise ValueError
        """
        _responsibility_id = -1
        if ((a_role_id == Roles_Definition.GAKYU_COMMITEE) or
            (a_role_id == Roles_Definition.GYOJI_COMMITEE) or
            (a_role_id == Roles_Definition.TOUBAN_COMMITEE) or
            (a_role_id == Roles_Definition.UNDOKAI_COMMITEE) or 
            (a_role_id == Roles_Definition.UNEI_COMMITEE)
            ):
            _responsibility_id = ResponsibilityLevel.TOUBAN_EXEMPT.value
        elif ((a_role_id == Roles_Definition.TOSHO_COMMITEE) or
              (a_role_id == Roles_Definition.SAFETY_COMMITEE)
              ):
            _responsibility_id = ResponsibilityLevel.COMMITTEE.value
        elif (
            # Handling of photo clue might be still NOT lucid as of 2024/08. 
            # Ref. https://groups.google.com/a/gjls.org/g/touban-group/c/8ikrmPQ15lk
            (a_role_id == Roles_Definition.PHOTO_CLUE) or
            (not a_role_id)):
            _responsibility_id = ResponsibilityLevel.GENERAL.value
        else:
            raise ValueError(f"""Obtained role '{a_role_id}' does NOT match any rule.
Make sure the input value conforms to the string expression this tool supports.
It is likely that the input values come all the way down from the input 'master' file (likely in '.xlsx' format).
So if that .xlsx file is an input data for your appliaction, check to see if there's any
                             """)
        return _responsibility_id
